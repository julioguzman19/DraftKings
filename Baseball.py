import pandas as pd
from openpyxl import load_workbook
from pulp import LpMaximize, LpProblem, LpStatus, lpSum, LpVariable

# Load the workbook
wb = load_workbook('DKSalaries.xlsx')

# ------------------------
# Select the active sheet
sheet = wb['Sheet1']  # replace 'Sheet1' with the name of your sheet

# Define the columns you want to keep
columns_to_keep = ['Position', 'Name', 'ID', 'Salary', 'AvgPointsPerGame']

# Get the max column count
max_column = sheet.max_column

# Traverse in reverse order (to avoid index change problem during deletion)
for i in range(max_column, 0, -1):
    # If the column header is not in columns_to_keep list, delete it
    if sheet.cell(row=1, column=i).value not in columns_to_keep:
        sheet.delete_cols(i)

# Add a new column header
sheet['G1'] = 'Predicted Points'  # Assuming 'G' is the next empty column

# Save the workbook
wb.save('DKSalaries.xlsx')

# Load the CSV file
df = pd.read_excel('DKSalaries.xlsx', sheet_name='Sheet1')  # Replace 'Sheet1' with your sheet name

# ------------------------
# Convert the Starting and Relief Pitchers into one
df['Position'] = df['Position'].replace({'RP': 'P', 'SP': 'P'})

# ------------------------
# Create the model
model = LpProblem(name="optimal-lineup", sense=LpMaximize)

# Create decision variables
player_vars = LpVariable.dicts("player", df.index, cat='Binary')

# Add objective function to the model
model += lpSum([df.loc[i, 'AvgPointsPerGame'] * player_vars[i] for i in df.index])

# Add salary constraint
model += lpSum([df.loc[i, 'Salary'] * player_vars[i] for i in df.index]) <= 50000

# Add position constraints
positions_needed = ['P', 'P', 'C', '1B', '2B', '3B', 'SS', 'OF', 'OF', 'OF']
for position in set(positions_needed):
    model += lpSum([player_vars[i] for i in df[df['Position'] == position].index]) == positions_needed.count(position)

# Add constraint so that each player can only be picked at most once
model += lpSum([player_vars[i] for i in df.index]) <= 10

# Add constraint for unique player IDs
selected_ids = set()
for i in df.index:
    player_id = df.loc[i, 'ID']
    if player_id in selected_ids:
        model += player_vars[i] == 0
    else:
        selected_ids.add(player_id)

# Solve the model
status = model.solve()

# Print the status of the solved LP
print(f"Status: {LpStatus[model.status]}")

# Get the players in the optimal lineup
lineup = df.iloc[[i for i in df.index if player_vars[i].value() == 1]]
lineup = lineup[lineup['Position'].isin(positions_needed)]
print(lineup)
