import pandas as pd
from openpyxl import load_workbook
from pulp import LpMaximize, LpProblem, LpStatus, lpSum, LpVariable
import requests
from bs4 import BeautifulSoup
import numpy as np
from pulp import LpProblem, LpMaximize, LpVariable, LpBinary, lpSum, LpStatus

#--------------------------------- Global Variables ---------------------------------
DK_CSV_NAME = 'DKSalaries.csv'
DK_EXCEL_NAME = 'DKSalaries.xlsx'
BASE_URL = "https://www.fantasypros.com/nfl/projections/{}.php?week=1"
POSITIONS = ['qb', 'rb', 'wr', 'te', 'dst']

# A dictionary to map each position to its relevant stats and indices
position_stats_mapping = {
    'qb': {
        'indices': [2, 3, 4, 6, 7],
        'columns': ['PassYards', 'PassTD', 'Interceptions', 'RushYards', 'RushTD']
    },
    'rb': {
        'indices': [1, 2, 3, 4, 5],
        'columns': ['RushYards', 'RushTD', 'Receptions', 'ReceiveYards', 'ReceiveTD']
    },
    'wr': {
        'indices': [0, 1, 2, 4],
        'columns': ['Receptions', 'ReceiveYards', 'ReceiveTD', 'RushYards']
    },
    'te': {
        'indices': [0, 1, 2],
        'columns': ['Receptions', 'ReceiveYards', 'ReceiveTD']
    },
    'dst': {
        'indices': [0, 1, 2, 6],
        'columns': ['Sacks', 'Interceptions', 'FumbleRecover', 'PointsAllowed']
    }
}

#--------------------------------- Clean Up DK Excel Data ---------------------------------
def clean_dk_excel_data(): 
    # Load your CSV
    df = pd.read_csv(DK_CSV_NAME)

    # Save as Excel
    df.to_excel(DK_EXCEL_NAME, index=False)

    # Load the workbook
    wb = load_workbook(DK_EXCEL_NAME)

    # ------------------------
    # Select the active sheet
    sheet = wb['Sheet1'] 

    # Define the columns you want to keep
    columns_to_keep = ['Position', 'Name', 'ID', 'Salary', 'TeamAbbrev']

    # Get the max column count
    max_column = sheet.max_column

    # Trim whitespace for all cells
    for row in range(1, sheet.max_row + 1):
        for col in range(1, max_column + 1):
            cell = sheet.cell(row=row, column=col)
            if cell.value and isinstance(cell.value, str):  # Check if cell has a value and it's a string
                cell.value = cell.value.rstrip()  # trim the right whitespace

    # Traverse in reverse order (to avoid index change problem during deletion)
    for i in range(max_column, 0, -1):
        # If the column header is not in columns_to_keep list, delete it
        if sheet.cell(row=1, column=i).value not in columns_to_keep:
            sheet.delete_cols(i)

    # Add new column headers
    sheet['F1'] = 'PredictedPts'  # Assuming 'E' is the next empty column
    sheet['G1'] = 'Include? (y/n)'

    # Fill all cells in column 'F' with 'n' except for the header
    max_row = sheet.max_row
    for row in range(2, max_row + 1):  # starting from 2 to skip the header
        sheet[f'G{row}'] = 'NA'

    # Save the workbook
    wb.save(DK_EXCEL_NAME)

#--------------------------------- Update players required or exclude from lineup ---------------------------------
def user_update_players():
    wb = load_workbook(DK_EXCEL_NAME)
    sheet = wb['Sheet1']

    while True:
        player_name = input("\nType the name of the player you want to update (or type 'done' to proceed): ").strip()
        if player_name.lower() == 'done':
            break
        found = False
        for row in range(2, sheet.max_row +1):  # start from 2 to skip the header
            if sheet[f'B{row}'].value == player_name:  # assuming names are in column B
                found = True
                current_value = sheet[f'G{row}'].value
                new_value = input(f"\nCurrent value for {player_name} is {current_value}. Include? (y/n): ").strip().lower()
                while new_value not in ['y', 'n']:
                    print("Invalid input. Please enter y or n or done.")
                    new_value = input(f"\nCurrent value for {player_name} is {current_value}. Include? (y/n): ").strip().lower()
                sheet[f'G{row}'].value = new_value
                break

        if not found:
            print(f"\nPlayer named '{player_name}' not found. Please check the name and try again.")

    wb.save(DK_EXCEL_NAME)
#--------------------------------- SCRAPE & IMPORT DATA --------------------------------
def scrape_position(position):
    url = BASE_URL.format(position)
    response = requests.get(url)
    soup = BeautifulSoup(response.content, 'html.parser')
    
    # Extract all player names based on your selector
    player_names = [element.text for element in soup.select('td.player-label a.player-name')]
    
    if position == 'dst':
        player_names = [name.split()[-1] for name in player_names]  # Retain only the last word if position is 'dst'

    # Extract all stats elements
    all_stats_elements = soup.select('td.center')

    all_players_data = []
    for idx, player_name in enumerate(player_names):
        stats_indices = position_stats_mapping[position]['indices']
        stats_elements_for_player = [all_stats_elements[i + (len(all_stats_elements) // len(player_names)) * idx] for i in stats_indices]
        stats = [float(stat_element.text) for stat_element in stats_elements_for_player]

        data = {column: stat for column, stat in zip(position_stats_mapping[position]['columns'], stats)}
        data["Player"] = player_name
        all_players_data.append(data)  # Append the dictionary to the list

    df = pd.DataFrame(all_players_data, columns=["Player"] + position_stats_mapping[position]['columns'])
    return df


#Import scraped data to new excel sheet 'Week1' in the existing Excel file DK_EXCEL_NAME.
def create_new_excel_sheet():
    columns_order = ['Player', 'PassYards', 'RushYards', 'ReceiveYards', 'PassTD', 
                     'RushTD', 'ReceiveTD', 'Receptions', 'Interceptions', 'Sacks', 
                     'FumbleRecover', 'PointsAllowed']
    
    # Create an empty dataframe with the desired structure
    consolidated_df = pd.DataFrame(columns=columns_order)

    # Loop over each position to scrape the data and append to the consolidated dataframe
    for position in POSITIONS:
        df_position = scrape_position(position)
        consolidated_df = pd.concat([consolidated_df, df_position], ignore_index=True)

    # Rearrange the dataframe columns to match the desired order
    consolidated_df = consolidated_df[columns_order]

    # Check and delete the "Week1" sheet if it already exists
    wb = load_workbook(DK_EXCEL_NAME)
    if "Week1" in wb.sheetnames:
        del wb["Week1"]

    # Open the existing Excel file and add a new sheet 'Week1' with the scraped data
    with pd.ExcelWriter(DK_EXCEL_NAME, engine='openpyxl', mode='a') as writer:
        # mode='a' means append mode, which ensures the existing content in the Excel file is not overwritten
        consolidated_df.to_excel(writer, sheet_name="Week1", index=False)

#--------------------------------- Merge Data Before Calculating Predicted Points ---------------------------------
def merge_data():
    # Load the Week1 and Sheet1 data
    df_week1 = pd.read_excel(DK_EXCEL_NAME, sheet_name='Week1')
    df_sheet1 = pd.read_excel(DK_EXCEL_NAME, sheet_name='Sheet1')

    # Strip whitespaces from the 'Name' column
    df_sheet1['Name'] = df_sheet1['Name'].str.strip()

    # Merge dataframes early on to get all required columns together for point calculation
    merged_df = df_sheet1.merge(df_week1[['Player', 'PassYards', 'RushYards', 'ReceiveYards', 'PassTD', 'RushTD', 'ReceiveTD', 'Receptions', 'Interceptions', 'Sacks', 'FumbleRecover', 'PointsAllowed']], left_on='Name', right_on='Player', how='left')

    # Filling NaN values with 0 for specified columns
    columns_to_fill = ['PassYards', 'RushYards', 'ReceiveYards', 'PassTD', 'RushTD', 'ReceiveTD', 'Receptions', 'Interceptions', 'Sacks', 'FumbleRecover', 'PointsAllowed']
    for col in columns_to_fill:
        merged_df[col].fillna(0, inplace=True)

    return df_sheet1, merged_df
#--------------------------------- Calculate Predicted Points ---------------------------------
def calculate_predicted_points(df_sheet1, merged_df):
    # Calculate points based on conditions
    merged_df['predicted_points'] = (
        merged_df['PassYards'] * 0.04 + 
        np.where(merged_df['PassYards'] >= 300, 3, 0) + 
        merged_df['RushYards'] * 0.1 +
        np.where(merged_df['RushYards'] >= 100, 3, 0) + 
        merged_df['ReceiveYards'] * 0.1 +
        np.where(merged_df['ReceiveYards'] >= 100, 3, 0) + 
        merged_df['PassTD'] * 4 +
        merged_df['RushTD'] * 6 +
        merged_df['ReceiveTD'] * 6 +
        merged_df['Receptions'] * 1 + 
        np.where(merged_df['Position'] == 'QB', merged_df['Interceptions'] * -1, merged_df['Interceptions'] * 2) + 
        merged_df['Sacks'] * 1 +
        merged_df['FumbleRecover'] * 2 +
        np.select(
            condlist=[
                (merged_df['PointsAllowed'] >= 1) & (merged_df['PointsAllowed'] <= 6), 
                (merged_df['PointsAllowed'] >= 7) & (merged_df['PointsAllowed'] <= 13), 
                (merged_df['PointsAllowed'] >= 14) & (merged_df['PointsAllowed'] <= 20),
                (merged_df['PointsAllowed'] >= 21) & (merged_df['PointsAllowed'] <= 27),
                (merged_df['PointsAllowed'] >= 28)
            ], 
            choicelist=[7, 4, 1, 0, -1],
            default=0
        )
    )

    df_sheet1['PredictedPts'] = merged_df['predicted_points']

    # Check and delete the "Week1" sheet if it already exists
    wb = load_workbook(DK_EXCEL_NAME)
    if "Sheet1" in wb.sheetnames:
        del wb["Sheet1"]
        wb.save(DK_EXCEL_NAME)

    # Use pd.ExcelWriter to overwrite Sheet1 while preserving other sheets
    with pd.ExcelWriter(DK_EXCEL_NAME, engine='openpyxl', mode='a') as writer:
        df_sheet1.to_excel(writer, sheet_name="Sheet1", index=False)   


#--------------------------------- Optimize Lineup ---------------------------------
def optimize_lineup():
    # Load the .xlsx file
    df = pd.read_excel(DK_EXCEL_NAME, sheet_name='Sheet1')  

    # Create the model
    model = LpProblem(name="optimal-lineup", sense=LpMaximize)

    # Create decision variables
    player_vars = LpVariable.dicts("player", df.index, cat='Binary')

    # Add objective function to the model
    model += lpSum([df.loc[i, 'PredictedPts'] * player_vars[i] for i in df.index])

    # Add salary constraint
    model += lpSum([df.loc[i, 'Salary'] * player_vars[i] for i in df.index]) <= 50000

    # Define the positions and their needed counts
    POSITIONS_NEEDED = {'QB': 1, 'RB': 2, 'WR': 3, 'TE': 1, 'DST': 1}
    INCREASE_POSITIONS = ['RB', 'WR', 'TE']

    # Create binary variables to denote if a position gets an additional player
    INCREASE_RB = LpVariable("INCREASE_RB", 0, 1, LpBinary)
    INCREASE_WR = LpVariable("INCREASE_WR", 0, 1, LpBinary)
    INCREASE_TE = LpVariable("INCREASE_TE", 0, 1, LpBinary)

    # Ensure only one position gets the additional player
    model += INCREASE_RB + INCREASE_WR + INCREASE_TE == 1

    # Constraints for positions that can have an additional player
    model += lpSum([player_vars[i] for i in df[df['Position'] == 'RB'].index]) == POSITIONS_NEEDED['RB'] + INCREASE_RB
    model += lpSum([player_vars[i] for i in df[df['Position'] == 'WR'].index]) == POSITIONS_NEEDED['WR'] + INCREASE_WR
    model += lpSum([player_vars[i] for i in df[df['Position'] == 'TE'].index]) == POSITIONS_NEEDED['TE'] + INCREASE_TE

    # Constraints for positions with a fixed count
    fixed_positions = [p for p in POSITIONS_NEEDED if p not in INCREASE_POSITIONS]
    for position in fixed_positions:
        model += lpSum([player_vars[i] for i in df[df['Position'] == position].index]) == POSITIONS_NEEDED[position]

    # Constraint for excluding or including specific players based on user choice
    for idx, row in df.iterrows():
        if row['Include? (y/n)'] == 'y':
            model += player_vars[idx] == 1
        elif row['Include? (y/n)'] == 'n':
            model += player_vars[idx] == 0

    # Add constraint for unique player IDs
    selected_ids = set()
    for i in df.index:
        player_id = df.loc[i, 'ID']
        if player_id in selected_ids:
            model += player_vars[i] == 0
        else:
            selected_ids.add(player_id)

    # Solve the model
    status = model.solve()

    # Print the status of the solved LP
    print(f"Status: {LpStatus[model.status]}")

    # Get the players in the optimal lineup
    lineup = df.iloc[[i for i in df.index if player_vars[i].value() == 1]]
    lineup = lineup[lineup['Position'].isin(POSITIONS_NEEDED)]
    print(lineup)
#--------------------------------- Call Helper Methods ---------------------------------
clean_dk_excel_data()
user_update_players()
create_new_excel_sheet()
df_sheet1, merged_df = merge_data()
calculate_predicted_points(df_sheet1, merged_df)
optimize_lineup()